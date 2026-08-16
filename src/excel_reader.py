from pathlib import Path
from openpyxl import load_workbook


def read_excel_files(input_folder):

    input_folder = Path(input_folder)

    supported_extensions = {
        ".xlsx",
        ".xlsm",
        ".xltx",
        ".xltm"
    }

    excel_files = [
        file_path
        for file_path in input_folder.iterdir()
        if (
            file_path.is_file()
            and file_path.suffix.lower()
            in supported_extensions
        )
    ]

    print(
        f"Excel files found: {len(excel_files)}"
    )

    all_records = []

    for file_path in excel_files:

        print(
            f"Reading: {file_path.name}"
        )

        workbook = load_workbook(
            file_path,
            data_only=True
        )

        sheet = workbook.active

        headers = [
            str(cell.value).strip().lower()
            for cell in sheet[1]
        ]

        for row in sheet.iter_rows(
            min_row=2,
            values_only=True
        ):

            record = dict(
                zip(headers, row)
            )

            all_records.append(
                record
            )

        workbook.close()

    return all_records