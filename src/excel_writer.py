from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


def create_workbook():

    workbook = Workbook()

    return workbook


def format_header(sheet):

    for cell in sheet[1]:

        cell.font = Font(
            bold=True
        )

        cell.fill = PatternFill(
            fill_type="solid",
            fgColor="1F4E78"
        )

        cell.alignment = Alignment(
            horizontal="center"
        )


def format_currency(sheet, columns):

    for column in columns:

        for cell in sheet[column][1:]:

            cell.number_format = '₹#,##0.00'


def format_dates(sheet):

    for cell in sheet["A"][1:]:

        cell.number_format = "yyyy-mm-dd"


def auto_adjust_columns(sheet):

    for column_cells in sheet.columns:

        max_length = 0

        column_letter = get_column_letter(
            column_cells[0].column
        )

        for cell in column_cells:

            if cell.value is not None:

                length = len(str(cell.value))

                if length > max_length:
                    max_length = length

        sheet.column_dimensions[
            column_letter
        ].width = max_length + 2


def freeze_header(sheet):

    sheet.freeze_panes = "A2"


def add_table(sheet):

    last_row = sheet.max_row
    last_column = sheet.max_column

    table_range = (
        f"A1:{get_column_letter(last_column)}{last_row}"
    )

    table = Table(
        displayName="SalesTable",
        ref=table_range
    )

    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )

    table.tableStyleInfo = style

    sheet.add_table(table)


def create_sales_sheet(workbook, data):

    sheet = workbook.active

    sheet.title = "Sales Data"

    headers = [
        "Date",
        "Employee",
        "Product",
        "Quantity",
        "Price",
        "Total"
    ]

    sheet.append(headers)

    for item in data:

        sheet.append([
            item["date"],
            item["employee"],
            item["product"],
            item["quantity"],
            item["price"],
            item["total"]
        ])

    format_header(sheet)

    format_dates(sheet)

    format_currency(
        sheet,
        ["E", "F"]
    )

    freeze_header(sheet)

    auto_adjust_columns(sheet)

    add_table(sheet)

    return sheet


def create_summary_sheet(workbook, summary):

    sheet = workbook.create_sheet(
        "Summary"
    )

    sheet.append([
        "Metric",
        "Value"
    ])

    rows = [
        ("Total Revenue", summary["total_revenue"]),
        ("Total Quantity", summary["total_quantity"]),
        ("Average Sale", summary["average_sale"]),
        ("Highest Sale", summary["highest_sale"]),
        ("Lowest Sale", summary["lowest_sale"]),
        ("Number of Employees", summary["number_of_employees"]),
        ("Number of Products", summary["number_of_products"]),
        ("Total Records", summary["total_records"])
    ]

    for row in rows:

        sheet.append(row)

    format_header(sheet)

    format_currency(
        sheet,
        ["B"]
    )

    freeze_header(sheet)

    auto_adjust_columns(sheet)

    return sheet


def create_employee_sheet(
    workbook,
    employee_data
):

    sheet = workbook.create_sheet(
        "Employee Summary"
    )

    sheet.append([
        "Employee",
        "Total Sales",
        "Total Quantity",
        "Orders"
    ])

    for employee, values in employee_data.items():

        sheet.append([
            employee,
            values["total_sales"],
            values["total_quantity"],
            values["number_of_orders"]
        ])

    format_header(sheet)

    format_currency(
        sheet,
        ["B"]
    )

    freeze_header(sheet)

    auto_adjust_columns(sheet)

    return sheet


def create_product_sheet(
    workbook,
    product_data
):

    sheet = workbook.create_sheet(
        "Product Summary"
    )

    sheet.append([
        "Product",
        "Total Sales",
        "Total Quantity",
        "Orders"
    ])

    for product, values in product_data.items():

        sheet.append([
            product,
            values["total_sales"],
            values["total_quantity"],
            values["number_of_orders"]
        ])

    format_header(sheet)

    format_currency(
        sheet,
        ["B"]
    )

    freeze_header(sheet)

    auto_adjust_columns(sheet)

    return sheet


def save_workbook(workbook, output_file):

    workbook.save(output_file)

    print(
        f"Excel report created: {output_file}"
    )